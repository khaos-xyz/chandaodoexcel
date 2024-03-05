# -*-coding: Utf-8 -*-
# @File : test3 .py
"""
# author: zhangyuxiong66
# Time：2024/1/31 15:16
# @Desc:
"""

import openpyxl
import pandas as pd
import matplotlib.pyplot as plt

# 读取 Excel 文件并创建副本
file_path = 'BUG统计日报--20240130.xlsx'
xl = openpyxl.load_workbook(file_path)
copy_xl = openpyxl.Workbook()
for sheet in xl.worksheets:
    copy_sheet = copy_xl.create_sheet(sheet.title)
    for row in sheet.iter_rows(values_only=True):
        copy_sheet.append(row)

# 转换副本文件中的公式单元格为值单元格
for sheet in copy_xl.worksheets:
    for row in sheet.iter_rows():
        for cell in row:
            if cell.data_type == 'f':
                cell.value = cell.value

# 设置中文字体
plt.rcParams['font.family'] = ['SimHei']

# 循环处理每个工作表
for sheet_name in copy_xl.sheetnames:
    ws = copy_xl[sheet_name]
    data = []
    merged_ranges = ws.merged_cells.ranges

    for row in ws.iter_rows(values_only=True):
        data.append(row)

    if not data:
        continue  # 跳过空数据

    cols = data[0]  # 第一行为列名
    data = data[1:]  # 去除第一行，保留数据部分

    # 处理合并单元格
    for merged_range in merged_ranges:
        start_row, start_col, end_row, end_col = merged_range.min_row, merged_range.min_col, merged_range.max_row, merged_range.max_col
        merged_value = data[start_row - 2][start_col - 1]  # 获取合并单元格的值
        for row in data[start_row - 1: end_row]:
            row[start_col - 1] = merged_value

    # 处理公式单元格
    for row in data:
        for i, value in enumerate(row):
            if isinstance(value, openpyxl.cell.cell.Cell) and value.data_type == 'f':
                row[i] = value.value

    df = pd.DataFrame(data, columns=cols)

    img_path = f'{sheet_name}.png'  # 图片保存路径
    plt.figure()
    plt.axis('off')
    plt.table(cellText=df.values, colLabels=df.columns, cellLoc='center', loc='center')
    plt.savefig(img_path, bbox_inches='tight', pad_inches=0)
    plt.close()

print("图片生成完成")













