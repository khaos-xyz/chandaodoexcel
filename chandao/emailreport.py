# -*-coding: Utf-8 -*-
# @File : emailreport .py
"""
# author: zhangyuxiong66
# Time：2024/1/25 20:28
# @Desc:
"""
import doexcel
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.image import MIMEImage
from email.mime.base import MIMEBase
from email import encoders
# from openpyxl import load_workbook
import openpyxl
import schedule
import time
import datetime
import os
from PIL import ImageGrab
import math
import pyautogui
import magic
from email.mime.application import MIMEApplication
import openpyxl
import pandas as pd
import matplotlib.pyplot as plt



# 发送方和接收方的信息
sender_email = "ceshi@jiuqingsk.com"
receiver_email = "zhangyx@jiuqingsk.com"
password = "jss3mPZnw7wFSigR"
today = time.strftime("%Y%m%d")


def send_email_with_screenshot():

    doexcel.doexcle()
    # 加载 Excel 文件
    excel_file = f"BUG统计日报--{today}.xlsx"
    # create_image(excel_file=excel_file)
    xl = openpyxl.load_workbook(excel_file)
    copy_xl = openpyxl.Workbook()
    for sheet in xl.worksheets:
        copy_sheet = copy_xl.create_sheet(sheet.title)
        for row in sheet.iter_rows(values_only=True):
            copy_sheet.append(row)

    # 转换副本文件中的公式单元格为值单元格
    for sheet in copy_xl.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if cell.data_type == 'f':
                    cell.value = cell.value

    # 设置中文字体
    plt.rcParams['font.family'] = ['SimHei']

    # 循环处理每个工作表
    for sheet_name in copy_xl.sheetnames:
        ws = copy_xl[sheet_name]
        data = []
        merged_ranges = ws.merged_cells.ranges

        for row in ws.iter_rows(values_only=True):
            data.append(row)

        if not data:
            continue  # 跳过空数据

        cols = data[0]  # 第一行为列名
        data = data[1:]  # 去除第一行，保留数据部分

        # 处理合并单元格
        for merged_range in merged_ranges:
            start_row, start_col, end_row, end_col = merged_range.min_row, merged_range.min_col, merged_range.max_row, merged_range.max_col
            merged_value = data[start_row - 2][start_col - 1]  # 获取合并单元格的值
            for row in data[start_row - 1: end_row]:
                row[start_col - 1] = merged_value

        # 处理公式单元格
        for row in data:
            for i, value in enumerate(row):
                if isinstance(value, openpyxl.cell.cell.Cell) and value.data_type == 'f':
                    row[i] = value.value

        df = pd.DataFrame(data, columns=cols)

        img_path = f'{sheet_name}.png'  # 图片保存路径
        plt.figure()
        plt.axis('off')
        plt.table(cellText=df.values, colLabels=df.columns, cellLoc='center', loc='center')
        plt.savefig(img_path, bbox_inches='tight', pad_inches=0)
        plt.close()

    print("图片生成完成!")
    # wb = openpyxl.load_workbook(excel_file)
    # ws = wb['Bug统计']
    # 获取第一个 sheet
    # worksheet = wb.active

    # 获取表格活动范围的位置和大小
    # left, top = ws.min_column, ws.min_row
    # width = (ws.max_column - ws.min_column + 1) * 20
    # height = (ws.max_row - ws.min_row + 1) * 15
    #
    # # 使用 pyautogui 截取屏幕截图
    # screenshot = pyautogui.screenshot(region=(left, top, width, height))
    #
    # # 将截图保存到本地文件
    # screenshot_file = 'screenshot.png'
    # screenshot.save(screenshot_file)

    # 创建邮件对象
    msg = MIMEMultipart()
    msg['From'] = sender_email
    msg['To'] = receiver_email
    msg['Subject'] = f"【Bug统计日报--{today}】"

    # 添加邮件正文（使用 HTML 格式）
    # body = '<html><body>今日Bug统计如下：<br><img src="cid:screenshot"></body></html>'
    # msg.attach(MIMEText(body, 'html'))



    # 将本地文件路径作为参数传递给 MIMEImage
    with open(img_path, 'rb') as f:
        image_data = f.read()
    image = MIMEImage(image_data, _subtype='png')
    image.add_header('Content-ID', '<screenshot>')
    msg.attach(image)

    # 添加 Excel 文件附件
    with open(excel_file, 'rb') as f:
        attach = MIMEApplication(f.read(), _subtype=magic.from_file(excel_file, mime=True))
        attach.add_header('Content-Disposition', 'attachment', filename=excel_file)
        msg.attach(attach)

    # 删除本地截图文件
    # os.remove(screenshot_file)

    # 发送邮件
    server = smtplib.SMTP('smtp.feishu.cn', 587)
    # smtp.set_debuglevel(2)
    server.starttls()
    server.login(sender_email, password)
    text = msg.as_string()
    server.sendmail(sender_email, receiver_email, text)
    server.quit()
    print("邮件发送完成!")

# 设置每天定时发送时间
# schedule.every().day.at("09:11").do(send_email_with_screenshot)

# while True:
#     schedule.run_pending()
#     time.sleep(1)
if __name__ == "__main__":

    send_email_with_screenshot()
